# Домашнее задание №11: Работа с комплексными файлами - excel, json, word. Git и Jira
# Выполните следующее задание:
# Задание №1
# а) Прочитайте из трёх excel файлов (заранее создайте их, внутри 1111, 2222, 3333).
# б) Отсортируйте полученную матрицу в порядке убывания.
# в) Запишите это в один файл, изменив шрифт и обернув в границы.

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Side

values = [1111, 2222, 3333]
file_names = ["a.xlsx", "b.xlsx", "c.xlsx"]

for name, value in zip(file_names, values):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = value
    wb.save(name)

matrix = []

for name in file_names:
    wb = load_workbook(name)
    ws = wb.active
    matrix.append(ws.cell(row=1, column=1).value)

matrix.sort(reverse=True)

wb = Workbook()
ws = wb.active

bold_font = Font(bold=True)
thin_border = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin')
)

for row, value in enumerate(matrix, start=1):
    cell = ws.cell(row=row, column=1)
    cell.value = value
    cell.font = bold_font
    cell.border = thin_border

wb.save("result.xlsx")
