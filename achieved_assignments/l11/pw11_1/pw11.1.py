# Практическая работа №11: Работа с комплексными файлами - excel, json, word. Git и Jira
# Выполните следующие задания:
# Задание №1
# а) Создайте excel файл в операционной системе, заполните его данными в одну строку,
# от 1 до 10. Во второй строке от а и до 10 буквы.
# б) Прочитайте эти две строки.
# в) Перезапишите их в другой файл, но вертикально, т.е. первая строка – первый столбец.

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

for col, value in enumerate(range(1, 11), start=1):
    ws.cell(row=1, column=col).value = value

for col, value in enumerate(range(ord('a'), ord('k')), start=1):
    ws.cell(row=2, column=col).value = chr(value)

wb.save("input.xlsx")

wb = Workbook()
ws_out = wb.active

numbers = [cell.value for cell in ws[1]]
letters = [cell.value for cell in ws[2]]

for row, value in enumerate(numbers, start=1):
    ws_out.cell(row=row, column=1).value = value

for row, value in enumerate(letters, start=1):
    ws_out.cell(row=row, column=2).value = value

wb.save("output.xlsx")
